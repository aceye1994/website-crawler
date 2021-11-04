#!/usr/bin/python
# -*- coding: utf-8 -*-

# File name: crawler_illness.py
# Author: Zhaowei Tan
# Contact: tanzw94@gmail.com

from bs4 import BeautifulSoup
import urllib2, bs4
from openpyxl import Workbook
import xlrd
import xlwt

# data=xlrd.open_workbook('department.xlsx')
workbook = Workbook()
sheet = workbook.create_sheet("doctor")
sheet.cell(row=1, column=1).value="doctor name"
sheet.cell(row=1, column=2).value="hospital"
# sheet.cell(row=1, column=3).value="province"
sheet.cell(row=1, column=3).value="department"
sheet.cell(row=1, column=4).value="level"
sheet.cell(row=1, column=5).value="speciality"
sheet.cell(row=1, column=6).value="total votes"
#sheet.cell(row=1, column=8).value="total votes 2"
#sheet.cell(row=1, column=9).value="webpage"

n_doctor = 0

def isElementClass(element, className):
  if type(element) == bs4.element.Tag and element.get('class') != None and e.get('class')[0] == className:
    return True
  else:
    return False

def process_page(soup, loc, url):
  global n_doctor
  doctor_list = soup.find_all("div", class_="oh zoom lh180")
  # print len(doctor_list)

  # Process speciality
  doctor_speciality = []
  p_list = soup.find_all("p")
  for p in p_list:
    if (p.encode('utf-8').find("擅长：") > 0):
      doctor_speciality.append(p.string[3:])
  if (len(doctor_list) != len(doctor_speciality)):
    print "find error: page ", url

  # print soup.prettify()
  for i in range(len(doctor_list)):
    doctor = doctor_list[i]
    
    doctor_info_list = doctor.contents
    # First row: name, level, hospital
    doctor_name = doctor_info_list[1].contents[0].string
    doctor_level = doctor_info_list[1].contents[2].string
    doctor_hospital = doctor_info_list[1].contents[4].string
    # print doctor_name, doctor_hospital, doctor_level
    
    # Third row: Voting
    voting_info = doctor_info_list[5].contents[1].string
    voting_pos = voting_info.find('/')
    voting_1 = voting_info[8:voting_pos]
    voting_2 = voting_info[voting_pos+6:]
    # print voting_info, voting_1, voting_2
    
    # print doctor_hospital
    y_pos = doctor_hospital.encode('utf-8').find("医院")
    y_pos_bkup = doctor_hospital.encode('utf-8').find("院")
    if (y_pos > 0):
      hospital_name = doctor_hospital.encode('utf-8')[0:y_pos+6]
      deparment_name = doctor_hospital.encode('utf-8')[y_pos+6:]
    elif y_pos_bkup > 0:
      hospital_name = doctor_hospital.encode('utf-8')[0:y_pos_bkup+3]
      deparment_name = doctor_hospital.encode('utf-8')[y_pos_bkup+3:]
    else:
      hospital_name = ""
      deparment_name = doctor_hospital

    # Write to excel
    n_doctor = n_doctor + 1
    sheet.cell(row = n_doctor + 1,column = 1).value=doctor_name
    sheet.cell(row = n_doctor + 1,column = 2).value=hospital_name
    # sheet.cell(row = n_doctor + 1,column = 3).value=loc
    sheet.cell(row = n_doctor + 1,column = 3).value=deparment_name
    sheet.cell(row = n_doctor + 1,column = 4).value=doctor_level
    sheet.cell(row = n_doctor + 1,column = 5).value=doctor_speciality[i]
    sheet.cell(row = n_doctor + 1,column = 6).value=voting_1
    # sheet.cell(row = n_doctor + 1,column = 8).value=voting_2
    # sheet.cell(row = n_doctor + 1,column = 9).value=url



# initial_page = "http://www.haodf.com/jibing/keluoenbing/daifu_all_all_all_all_all.htm"
initial_page = "https://www.haodf.com/jibing/hongbanlangchuang/daifu_all_all_all_all_all_all_1.htm"

opener = urllib2.build_opener()
opener.addheaders = [('User-agent', 'Mozilla/5.0')]
f = opener.open(initial_page)
content = f.read()
soup = BeautifulSoup(content, "html.parser", from_encoding="gbk")
# print soup.prettify()

ct_list = soup.find_all("ul",class_="clearfix area_box_list")
ct = ct_list[0]
# print type(ct)
# print ct.prettify()
for child in ct.find_all('a'):
  loc_name = child.string
  print loc_name.encode('utf-8')
  page_href =  "http:" + child['href']

  print page_href
  if page_href == initial_page:
    continue

  # print "processing: page 1 " + loc_name
  f = opener.open(page_href)
  content = f.read()
  soup_new = BeautifulSoup(content, "html.parser", from_encoding="gbk")
  
  try: # it's possible to search no result under the homepage
    total_pages = soup_new.find("font", class_="black pl5 pr5")
    # print total_pages.string
    total_pages = int(total_pages.string)
  except Exception, e:
    total_pages = 0

  process_page(soup_new, loc_name, page_href)
  # print soup.prettify()

  for i in range(2, 17):
  # for i in range(2, total_pages + 1):
    new_page_href = page_href[:-5] + "_" + str(i) + ".htm"
    # print new_page_href
    # print "processing: page " + str(i) + " " + loc_name
    try:
      f = opener.open(new_page_href)
      content = f.read()
      soup_new = BeautifulSoup(content, "html.parser", from_encoding="gbk")
      process_page(soup_new, loc_name, new_page_href)
    except Exception, e:
      pass
  break

workbook.save('doctor_hongbanlangchuang.xlsx')
