#!/usr/bin/python
# -*- coding: utf-8 -*- 

# File name: crawler_illness_hospital.py
# Author: Zhaowei Tan
# Contact: tanzw94@gmail.com

from bs4 import BeautifulSoup
import urllib2, bs4
from openpyxl import Workbook
import xlrd
import xlwt

# data=xlrd.open_workbook('department.xlsx')
workbook = Workbook()
sheet = workbook.create_sheet("hospital")
sheet.cell(row=1, column=1).value="hospital"
sheet.cell(row=1, column=2).value="province"
sheet.cell(row=1, column=3).value="city"
sheet.cell(row=1, column=4).value="level"
sheet.cell(row=1, column=5).value="# of doctors"
sheet.cell(row=1, column=6).value="votings"
sheet.cell(row=1, column=7).value="url"

n_hospital = 0

def process_page(soup, loc, url):
  global n_hospital
  hospital_list = soup.find_all("tr", class_="con_list")

  for hospital in hospital_list:
    
    # process hospital name
    hospital_name = hospital.contents[1].a.string
    # process hospital city
    hospital_city = hospital.contents[3].string
    # process hospital level
    hospital_level = hospital.contents[5].string
    # process hospital number of doctors
    hospital_ndoctor = hospital.contents[7].span.string
    # process hospital votes
    hospital_votes = hospital.contents[9].span.string

    # Write to excel
    n_hospital = n_hospital + 1
    sheet.cell(row = n_hospital + 1,column = 1).value=hospital_name
    sheet.cell(row = n_hospital + 1,column = 2).value=loc
    sheet.cell(row = n_hospital + 1,column = 3).value=hospital_city
    sheet.cell(row = n_hospital + 1,column = 4).value=hospital_level
    sheet.cell(row = n_hospital + 1,column = 5).value=hospital_ndoctor
    sheet.cell(row = n_hospital + 1,column = 6).value=hospital_votes
    sheet.cell(row = n_hospital + 1,column = 7).value=url


# initial_page = "http://www.haodf.com/jibing/keluoenbing_yiyuan_all_all_all_all_1.htm"
initial_page = "https://www.haodf.com/jibing/leifengshixingguanjieyan_yiyuan_all_all_all_all_1.htm"

opener = urllib2.build_opener()
opener.addheaders = [('User-agent', 'Mozilla/5.0')]
f = opener.open(initial_page)
content = f.read()
soup = BeautifulSoup(content, "html.parser", from_encoding="gbk")
# print soup.prettify()

ct_list = soup.find_all("ul",class_="clearfix area_box_list")
ct = ct_list[0]
# print type(ct)
# print ct.prettify()
for child in ct.find_all('a'):
  loc_name = child.string
  page_href = "http:" + child['href']
  
  print page_href

  if page_href == initial_page:
    continue

  #print "processing: page " + loc_name
  
  try: # handle non-existent province page
    f = opener.open(page_href)
    content = f.read()
    soup_new = BeautifulSoup(content, "html.parser", from_encoding="gbk")
  except Exception, e:
    continue

  try: # Process pagination pages -- currently not needed
    total_pages = soup_new.find("font", class_="black pl5 pr5")
    # print total_pages.string
    total_pages = int(total_pages.string)
  except Exception, e:
    total_pages = 0

  process_page(soup_new, loc_name, page_href)
  # print soup.prettify()


  # Process pagination pages -- currently not needed
  for i in range(2, total_pages + 1):
    new_page_href = page_href[:-6] + "_" + str(i) + ".htm"
    print new_page_href
    f = opener.open(new_page_href)
    content = f.read()
    soup_new = BeautifulSoup(content, "html.parser", from_encoding="gbk")
    process_page(soup_new, loc_name, new_page_href)
  
  # break

workbook.save('hospital_leifengshixingguanjieyan.xlsx')
