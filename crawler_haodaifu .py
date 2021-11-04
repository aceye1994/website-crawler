from bs4 import BeautifulSoup
import urllib2, bs4
from openpyxl import Workbook
import xlrd
import xlwt

# data=xlrd.open_workbook('department.xlsx')
workbook = Workbook()
sheet_first = workbook.create_sheet("jiangsu")
sheet_first.cell(row=1, column=1).value="hospital name"
sheet_first.cell(row=1, column=2).value="hospital class"
sheet_first.cell(row=1, column=3).value="province"
sheet_first.cell(row=1, column=4).value="city"
sheet_first.cell(row=1, column=5).value="hospital website"
# sheet_first.write(0,1,'hospital class')
# sheet_first.write(0,2,'province')
# sheet_first.write(0,3,'city')
# sheet_first.write(0,4,'hospital website')

def isElementClass(element, className):
  if type(element) == bs4.element.Tag and element.get('class') != None and e.get('class')[0] == className:
    return True
  else:
    return False

hospitals=[]
n_hospital = 0
url_group = ["http://www.haodf.com/yiyuan/beijing/list.htm","http://www.haodf.com/yiyuan/shanghai/list.htm","http://www.haodf.com/yiyuan/guangdong/list.htm","http://www.haodf.com/yiyuan/guangxi/list.htm", \
"http://www.haodf.com/yiyuan/jiangsu/list.htm","http://www.haodf.com/yiyuan/zhejiang/list.htm","http://www.haodf.com/yiyuan/anhui/list.htm","http://www.haodf.com/yiyuan/jiangxi/list.htm", \
"http://www.haodf.com/yiyuan/fujian/list.htm","http://www.haodf.com/yiyuan/shandong/list.htm","http://www.haodf.com/yiyuan/sx/list.htm", \
"http://www.haodf.com/yiyuan/hebei/list.htm","http://www.haodf.com/yiyuan/henan/list.htm","http://www.haodf.com/yiyuan/tianjin/list.htm", \
"http://www.haodf.com/yiyuan/liaoning/list.htm","http://www.haodf.com/yiyuan/heilongjiang/list.htm","http://www.haodf.com/yiyuan/shanghai/list.htm", \
"http://www.haodf.com/yiyuan/jilin/list.htm","http://www.haodf.com/yiyuan/hubei/list.htm","http://www.haodf.com/yiyuan/hunan/list.htm","http://www.haodf.com/yiyuan/sichuan/list.htm", \
"http://www.haodf.com/yiyuan/chongqing/list.htm","http://www.haodf.com/yiyuan/shanxi/list.htm","http://www.haodf.com/yiyuan/gansu/list.htm", \
"http://www.haodf.com/yiyuan/yunnan/list.htm","http://www.haodf.com/yiyuan/xinjiang/list.htm","http://www.haodf.com/yiyuan/neimenggu/list.htm", \
"http://www.haodf.com/yiyuan/hainan/list.htm","http://www.haodf.com/yiyuan/guizhou/list.htm","http://www.haodf.com/yiyuan/qinghai/list.htm", \
"http://www.haodf.com/yiyuan/ningxia/list.htm","http://www.haodf.com/yiyuan/xizang/list.htm"]
for url in url_group:
  opener = urllib2.build_opener()
  opener.addheaders = [('User-agent', 'Mozilla/5.0')]
  f = opener.open(url)
  content = f.read()
  soup = BeautifulSoup(content, "html.parser", from_encoding="gbk")
  ct_list = soup.find_all("div",class_="ct")
  ct = ct_list[1]
  for child in ct.children:
    if type(child) == bs4.element.Tag:
      # print child.get('class')
      if child.get('class') != None and child.get('class')[0] == 'm_title_green':
        # print child.string.encode('utf-8')
        city = child.string
        pass
      if child.get('class') != None and child.get('class')[0] == 'm_ctt_green':
        for li in child.find_all('li'):
          hospitals.append("http://www.haodf.com"+li.a.get('href'))
          n_hospital += 1
          sheet_first.cell(row=n_hospital+1,column=3).value=url
          # sheet_first.write(n_hospital,3,city)
          sheet_first.cell(row=n_hospital+1,column=4).value=city
          # sheet_first.write(n_hospital,4,hospitals[n_hospital-1])
          sheet_first.cell(row=n_hospital+1,column=5).value=hospitals[n_hospital-1]
          # print li.a.get('href'), li.a.string.encode('utf-8')
          # for span in li.find_all('span'):
          # print span.string.encode('utf-8')

try:
  hospital_names = []
  hospital_classes = []
  hospital_departments = []
  n_hospital = 0
  dic_department = {}
  list_department=[]
  list_sheet=[]
  # start to crawl each hospital
  for hospital in hospitals:
    try:
      n_hospital += 1
      print "Crawling website:", hospital
      opener_hospital = urllib2.build_opener()
      opener_hospital.addheaders = [('User-agent', 'Mozilla/5.0')]
      f_hospital = opener_hospital.open(hospital)
      content_hospital = f_hospital.read()
      soup_hospital = BeautifulSoup(content_hospital, "html.parser", from_encoding="gbk") 
      # print soup_hospital.prettify("gbk").decode('gbk').encode('utf-8')
      
      name_hospitals = soup_hospital.find_all("div",id="ltb")
      name_hospital = name_hospitals[0]
      # print depart_hospital.span.a.string.encode('utf-8')
      hospital_names.append(name_hospital.span.a.string.encode('utf-8'))
      hospital_name = name_hospital.span.a.string
      # sheet_first.write(n_hospital,0,hospital_name)
      sheet_first.cell(row=n_hospital+1,column=1).value=hospital_name


      try:
        class_hospitals=soup_hospital.find_all("div",class_="toptr")
        # print class_hospitals
        class_hospital=class_hospitals[0]
        # print class_hospital.p.a.string.encode('utf-8')
        # print class_hospital.p.contents[2].encode('utf-8')
        hospital_classes.append(class_hospital.p.contents[2].encode('utf-8'))
        hospital_class = class_hospital.p.contents[2]
        tmp = hospital_class.find(")")
        hospital_class = hospital_class[1:tmp]
        # sheet_first.write(n_hospital,1,hospital_class)
        sheet_first.cell(row=n_hospital+1,column=2).value=hospital_class
      except Exception, e:
        pass
      

      hospital_dename=[]
      hospital_depeople=[]
      depart_hospitals=soup_hospital.find_all("td",class_="font14")
      # print depart_hospitals
      for i in range(len(depart_hospitals)):
        depart_hospital = depart_hospitals[i]
        if not depart_hospital.string in dic_department:
          dic_department[depart_hospital.string] = []
          list_department.append(depart_hospital.string)
          sheet = workbook.create_sheet(depart_hospital.string)
          list_sheet.append(sheet)
        tmp2 = list_department.index(depart_hospital.string)
        # print
        # print depart_hospital.string.encode('utf-8')
        # print " ----- "
        for e in depart_hospital.next_elements:
          if (isElementClass(e, 'blue')):
            # print e.string.encode('utf-8')
            hospital_dename.append(e.string)
            if not e.string in dic_department[depart_hospital.string]:
              dic_department[depart_hospital.string].append(e.string)
              list_sheet[tmp2].cell(row=1,column=len(dic_department[depart_hospital.string])+1).value=e.string
            tmp = dic_department[depart_hospital.string].index(e.string)
          if (isElementClass(e, 'gray')):
            # print e.get('title').encode('utf-8').split()
            title_string = e.get('title').encode('utf-8')
            # print title_string
            start = title_string.find('\x9c')
            end = title_string.find('\xe4')
            # print "total doctors:", int(title_string[6 : end])
            hospital_depeople.append(title_string[6 : end])
            hospital_doctor = int(title_string[6 : end])
            list_sheet[tmp2].cell(row=n_hospital+1,column=tmp+2).value=hospital_doctor
          if (i != len(depart_hospitals) - 1 and e == depart_hospitals[i + 1]):
            break
          if (i == len(depart_hospitals) - 1 and isElementClass(e, 'textrt')):
            break
      # if n_hospital == 2:
      #   break
    except Exception, e:
      pass
finally:
  workbook.save('department_jiangsu.xlsx')
  print 'done'